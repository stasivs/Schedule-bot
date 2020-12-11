import os
import json

from openpyxl import load_workbook


def convert_directory(dir):
    merged_dict = dict()
    for file in os.listdir(dir):
        try:
            merged_dict.update(convert_file(os.path.join(dir, file)))
        except Exception:
            print(file)
    return merged_dict


def convert_file(file):
    wb = load_workbook(file)
    ws = wb.active

    day_of_the_weeks = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]
    groups = dict()
    columns = list(ws.iter_cols(min_row=3))

    # Размерживаю ячейки и заполняю их данными
    for items in sorted(ws.merged_cell_ranges):
        items = str(items)
        ws.unmerge_cells(items)

        temp = None
        for row in ws[items]:
            for cell in row:
                if cell.value:
                    temp = cell.value

        for row in ws[items]:
            for cell in row:
                cell.value = temp

    # Пробежка по всем столбцам
    for col in range(len(columns)):
        # Если не имеет названия группы, то проигнорировать столбец
        if not columns[col][0].value:
            continue

        # Инициализация словаря с расписанием на неделю (с учетом четности)
        daily_schedule = dict()

        cells = [[], []]
        counter = 0
        start = 2
        if "уч.нед." in str(columns[col][2].value):
            start += 1

        for cell in range(start, start + 60):
            counter += 1
            # Если неделя четная
            if cell % 2 == 0:
                cells[0].append(columns[col][cell].value)
            # Если неделя нечетная
            else:
                cells[1].append(columns[col][cell].value)

            if counter % 12 == 0:
                daily_schedule.update({day_of_the_weeks[counter // 12 - 1]: cells})
                cells = [[], []]

        # Немного видоизменим заголовок
        group = columns[col][0].value.lower().split()
        title = group[0] + " " + group[1] + " " + str(int(group[3]))

        if title in groups:
            groups[title].append(daily_schedule)
        else:
            groups[title] = [daily_schedule]

    return groups


if __name__ == "__main__":
    # res = convert_file("Schedule/GR 2 41-43.xlsx")
    # res = convert_file("Schedule/ISA 2 1-20.xlsx")
    res = convert_directory("Schedule")

    with open("data.json", "w", encoding="utf8") as file:
        json.dump(res, file)

